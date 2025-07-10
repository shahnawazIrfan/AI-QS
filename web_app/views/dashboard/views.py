from datetime import datetime, timedelta
from decimal import Decimal
import json
import logging
import os
import re
import boto3
from django.http import JsonResponse
import openpyxl
from core.portal_base import ViewBase
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import base64
from io import BytesIO
import matplotlib
from bson import ObjectId
from django.forms.models import model_to_dict
from django.db.models import Sum
from django.conf import settings
from openpyxl.styles import Font
from web_app import models
from django.db.models import Q
from django.utils import timezone
from django.db import transaction


matplotlib.use('Agg')

logger = logging.getLogger(__name__)


class UserDashboardView(ViewBase):
    TEMPLATE_NAME = 'dashboard/user_dashboard.html'

    def get(self, request, *args, **kwargs):

        context = {}

        return self.render(context)
    

class ClientProjectListingView(ViewBase):
    TEMPLATE_NAME = 'dashboard/client_project_listing.html'

    def get(self, request, *args, **kwargs):

        context = {}

        return self.render(context)
    

class ClientProjectDetailView(ViewBase):
    TEMPLATE_NAME = 'dashboard/client_project_details.html'

    def get(self, request, *args, **kwargs):

        context = {}

        return self.render(context)
    

class CostDashboardView(ViewBase):
    TEMPLATE_NAME = 'dashboard/cost_dashboard.html'

    def get(self, request, *args, **kwargs):
        # top card counts
        cost_summary = models.CostSummary.objects.all()
        total_contract_sum = round(sum(Decimal(str(obj.contract_sum)) for obj in cost_summary if obj.contract_sum), 2)
        certified_payments_sum = round(sum(Decimal(str(obj.certified_payments)) for obj in cost_summary if obj.certified_payments), 2)
        anticipated_payments_sum = round(sum(Decimal(str(obj.accrued_payments)) for obj in cost_summary if obj.accrued_payments), 2)
        forecast_expenditures_sum = round(sum(Decimal(str(obj.total_expenditure)) for obj in cost_summary if obj.total_expenditure), 2)
        total_variance_sum = round(sum(Decimal(str(obj.variance_total)) for obj in cost_summary if obj.variance_total), 2)
        total_variance_period_sum = round(sum(Decimal(str(obj.variance_period)) for obj in cost_summary if obj.variance_period), 2)

        # cost summary table
        total_new_cost_summary_sections = models.CostSummarySection.objects.count()
        costSummarySections = models.CostSummarySection.objects.prefetch_related("cost_summaries").all()

        new_cost_summary_data = []

        for section in costSummarySections:
            section_dict = {
                "section_id": section._id,
                "section_name": section.name,
                "rows": []
            }

            for cost_summary in section.cost_summaries.all():
                cost_summary_dict = model_to_dict(cost_summary, exclude=["section"])
                cost_summary_dict["id"] = str(cost_summary._id)
                cost_summary_dict["updated_at"] = timezone.localtime(cost_summary.updated_at).strftime("%m/%d/%Y")
                section_dict["rows"].append(cost_summary_dict)

            new_cost_summary_data.append(section_dict)


        # contract sum table
        total_new_contract_sum_sections = models.ContractSumSection.objects.count()
        ContractSumSections = models.ContractSumSection.objects.prefetch_related("contract_sums").all()

        new_contract_sum_data = []

        for section in ContractSumSections:
            section_dict = {
                "section_id": section._id,
                "section_name": section.name,
                "rows": []
            }

            for contract_sum in section.contract_sums.all():
                contract_sum_dict = model_to_dict(contract_sum, exclude=["section"])
                contract_sum_dict["id"] = str(contract_sum._id)
                contract_sum_dict["updated_at"] = timezone.localtime(contract_sum.updated_at).strftime("%m/%d/%Y")
                section_dict["rows"].append(contract_sum_dict)

            new_contract_sum_data.append(section_dict)


        # change breakdown table
        total_new_change_breakdown_sections = models.ChangeBreakDownSection.objects.count()
        ChangeBreakDownSections = models.ChangeBreakDownSection.objects.prefetch_related("change_breakdown").all()

        new_change_breakdown_data = []

        for section in ChangeBreakDownSections:
            section_dict = {
                "section_id": section._id,
                "section_name": section.name,
                "rows": []
            }

            for change_breakdown in section.change_breakdown.all():
                change_breakdown_dict = model_to_dict(change_breakdown, exclude=["section"])
                change_breakdown_dict["id"] = str(change_breakdown._id)
                change_breakdown_dict["updated_at"] = timezone.localtime(change_breakdown.updated_at).strftime("%m/%d/%Y")
                section_dict["rows"].append(change_breakdown_dict)

            new_change_breakdown_data.append(section_dict)


        # cost reporting graph sum values

        cost_reporting = models.CostReporting.objects.all()
        total_forecast_monthly = round(sum(Decimal(str(obj.forecast_monthly)) for obj in cost_reporting if obj.forecast_monthly), 2)
        total_actual_monthly = round(sum(Decimal(str(obj.actual_monthly)) for obj in cost_reporting if obj.actual_monthly), 2)
        total_forecast_comulative = round(sum(Decimal(str(obj.forecast_comulative)) for obj in cost_reporting if obj.forecast_comulative), 2)
        total_actual_comulative = round(sum(Decimal(str(obj.actual_comulative)) for obj in cost_reporting if obj.actual_comulative), 2)
        
        # cost reporting graph month values

        months = models.CostReporting.objects.values("month").distinct().order_by("month")
        months = [m["month"] for m in months if m["month"]]
        months = [date.strftime("%b %y") for date in months]

        context = {
            'new_cost_summary_data': new_cost_summary_data,
            'total_new_cost_summary_sections': total_new_cost_summary_sections,
            'new_contract_sum_data': new_contract_sum_data,
            'total_new_contract_sum_sections': total_new_contract_sum_sections,
            'new_change_breakdown_data': new_change_breakdown_data,
            'total_new_change_breakdown_sections': total_new_change_breakdown_sections,
            'total_contract_sum': total_contract_sum,
            'certified_payments_sum': certified_payments_sum,
            'anticipated_payments_sum': anticipated_payments_sum,
            'forecast_expenditures_sum': forecast_expenditures_sum,
            'total_variance_sum': total_variance_sum,
            'total_variance_period_sum': total_variance_period_sum,
            'total_forecast_monthly': total_forecast_monthly,
            'total_actual_monthly': total_actual_monthly,
            'total_forecast_comulative': total_forecast_comulative,
            'total_actual_comulative': total_actual_comulative,
            'months': months
        }

        return self.render(context)
    
    def post(self, request, *args, **kwargs):
        if "file" not in request.FILES:
            return JsonResponse({"error": "No file uploaded"}, status=400)

        file = request.FILES["file"]
        wb = openpyxl.load_workbook(file, data_only=True)

        try:
            with transaction.atomic():

                # === NEW COST SUMMARY ===
                if "NEW Cost Summary" not in wb.sheetnames:
                    return JsonResponse({"error": "Sheet 'NEW Cost Summary' not found"}, status=400)

                sheet = wb["NEW Cost Summary"]
                headers = [cell.value for cell in sheet[1]]
                required = {"ICMS2", "Item", "CONTRACT SUM", "CERTIFIED PAYMENTS TO CONTRACTOR", "ACCRUED & ANTICIPATED PAYMENTS", "TOTAL FORECAST EXPENDITURE", "VARIANCE - TOTAL ", "VARIANCE - IN PERIOD"}
                if missing := required - set(headers):
                    return JsonResponse({"error": f"Missing columns in 'NEW Cost Summary': {', '.join(missing)}"}, status=400)

                cs_sections, cs_items = [], []
                cs_section_id = None
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data = {
                        headers[i]: str(row[i]) if headers[i] == "ICMS2" else (row[i] if row[i] not in [None, ""] else 0.0)
                        for i in range(len(headers))
                    }
                    if "." not in str(data["ICMS2"]) and data["Item"]:
                        section = models.CostSummarySection(_id=str(ObjectId()), name=data["Item"])
                        cs_sections.append(section)
                        cs_section_id = section._id
                    elif "." in str(data["ICMS2"]) and data["Item"]:
                        cs_items.append(models.CostSummary(
                            _id=str(ObjectId()),
                            ref=data["ICMS2"],
                            item=data["Item"],
                            contract_sum=data["CONTRACT SUM"],
                            certified_payments=data["CERTIFIED PAYMENTS TO CONTRACTOR"],
                            accrued_payments=data["ACCRUED & ANTICIPATED PAYMENTS"],
                            total_expenditure=data["TOTAL FORECAST EXPENDITURE"],
                            variance_total=data["VARIANCE - TOTAL "],
                            variance_period=data["VARIANCE - IN PERIOD"],
                            section_id=cs_section_id
                        ))

                models.CostSummarySection.objects.bulk_create(cs_sections)
                models.CostSummary.objects.bulk_create(cs_items)

                # === NEW CONTRACT SUM ===
                if "NEW Contract Sum" not in wb.sheetnames:
                    return JsonResponse({"error": "Sheet 'NEW Contract Sum' not found"}, status=400)

                sheet = wb["NEW Contract Sum"]
                headers = [cell.value for cell in sheet[1]]
                required = {"ICMS2", "Item", "CONTRACT SUM", "CERTIFIED PAYMENTS TO CONTRACTOR"}
                if missing := required - set(headers):
                    return JsonResponse({"error": f"Missing columns in 'NEW Contract Sum': {', '.join(missing)}"}, status=400)

                csum_sections, csum_items = [], []
                csum_section_id = None
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data = {
                        headers[i]: str(row[i]) if headers[i] == "ICMS2" else (row[i] if row[i] not in [None, ""] else 0.0)
                        for i in range(len(headers))
                    }
                    if "." not in str(data["ICMS2"]) and data["Item"]:
                        section = models.ContractSumSection(_id=str(ObjectId()), name=data["Item"])
                        csum_sections.append(section)
                        csum_section_id = section._id
                    elif "." in str(data["ICMS2"]) and data["Item"]:
                        csum_items.append(models.ContractSum(
                            _id=str(ObjectId()),
                            ref=data["ICMS2"],
                            item=data["Item"],
                            contract_sum=data["CONTRACT SUM"],
                            certified_payments=data["CERTIFIED PAYMENTS TO CONTRACTOR"],
                            section_id=csum_section_id
                        ))

                models.ContractSumSection.objects.bulk_create(csum_sections)
                models.ContractSum.objects.bulk_create(csum_items)

                # === NEW EAChange ===
                if "NEW EAChange" not in wb.sheetnames:
                    return JsonResponse({"error": "Sheet 'NEW EAChange' not found"}, status=400)

                sheet = wb["NEW EAChange"]
                headers = [cell.value for cell in sheet[1]]
                required = {"Ref", "Item", "CONTRACT SUM", "CERTIFIED PAYMENTS TO CONTRACTOR", "ACCRUED & ANTICIPATED PAYMENTS", "TOTAL FORECAST EXPENDITURE", "VARIANCE - TOTAL ", "VARIANCE - IN PERIOD"}
                if missing := required - set(headers):
                    return JsonResponse({"error": f"Missing columns in 'NEW EAChange': {', '.join(missing)}"}, status=400)

                ea_sections, ea_items = [], []
                ea_section_id = None
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    data = {
                        headers[i]: str(row[i]) if headers[i] == "Ref" else (row[i] if row[i] not in [None, ""] else 0.0)
                        for i in range(len(headers))
                    }
                    if "." not in str(data["Ref"]) and data["Item"]:
                        section = models.ChangeBreakDownSection(_id=str(ObjectId()), name=data["Item"])
                        ea_sections.append(section)
                        ea_section_id = section._id
                    elif "." in str(data["Ref"]) and data["Item"]:
                        ea_items.append(models.ChangeBreakDown(
                            _id=str(ObjectId()),
                            ref=data["Ref"],
                            item=data["Item"],
                            certified_payments=data["CERTIFIED PAYMENTS TO CONTRACTOR"],
                            total_expenditure=data["TOTAL FORECAST EXPENDITURE"],
                            variance_total=data["VARIANCE - TOTAL "],
                            variance_period=data["VARIANCE - IN PERIOD"],
                            section_id=ea_section_id
                        ))

                models.ChangeBreakDownSection.objects.bulk_create(ea_sections)
                models.ChangeBreakDown.objects.bulk_create(ea_items)

            # === NEW COST REPORTING ===
            if "New Cost Reporting " not in wb.sheetnames:
                return JsonResponse({"error": "Sheet 'New Cost Reporting' not found"}, status=400)

            new_cost_reporting_sheet = wb["New Cost Reporting "]

            new_cost_reporting_required_columns = {"Interim Payments", "Month", "Forecast Monthly", "Actual Monthly", "Forecast Cumulative", "Actual Cumulative"}

            new_cost_reporting_headers = [cell.value for cell in new_cost_reporting_sheet[1]]

            new_cost_reporting_missing_columns = new_cost_reporting_required_columns - set(new_cost_reporting_headers)
            if new_cost_reporting_missing_columns:
                return JsonResponse(
                    {"error": f"Invalid format for Sheet 'New Cost Reporting': Missing columns - {', '.join(new_cost_reporting_missing_columns)}"},
                    status=400
                )

            for row in new_cost_reporting_sheet.iter_rows(min_row=2, values_only=True):
                item = {
                    new_cost_reporting_headers[i]: str(row[i]) if isinstance(row[i], (int, float)) 
                    else (row[i] if row[i] not in [None, "", "#REF!"] else 0.0)
                    for i in range(len(new_cost_reporting_headers))
                }

                try:
                    models.CostReporting.objects.create(
                        _id=str(ObjectId()),
                        interim_payments=item['Interim Payments'],
                        month=item['Month'] + timedelta(days=1),
                        forecast_monthly=item['Forecast Monthly'],
                        actual_monthly=item['Actual Monthly'],
                        forecast_comulative=item['Forecast Cumulative'],
                        actual_comulative=item['Actual Cumulative']
                    )
                except Exception:
                    pass

            return JsonResponse({"message": "File processed successfully"}, status=201)

        except Exception as e:
            return JsonResponse({"error": f"Processing failed: {str(e)}"}, status=500)
    

class RiskDashboardView(ViewBase):
    TEMPLATE_NAME = 'dashboard/risk_dashboard.html'

    def get(self, request, *args, **kwargs):

        # top card counts
        risk_register = models.RiskRegister.objects.all()
        total_risk_register = risk_register.count()
        total_mitigated_risk = risk_register.filter(mitigated_rating__isnull=False).count()
        total_unmitigated_risk = risk_register.filter(mitigated_rating__isnull=True, rating__isnull=False).count()
        total_design_development_risk = risk_register.filter(section__name__icontains="Design development risks").count()
        total_construction_risk = risk_register.filter(section__name__icontains="Construction risks").count()
        total_employer_change_risk = risk_register.filter(section__name__icontains="Employer change risks").count()
        total_employer_other_risk = risk_register.filter(Q(section__name__icontains="Project brief") | Q(section__name__icontains="Timescales") | Q(section__name__icontains="Financial") | Q(section__name__icontains="Management") | Q(section__name__icontains="Third party") | Q(section__name__icontains="Other")).count()

        # pie charts
        pie_charts = {}
        pie_charts_values = {}

        for i in range(5):
            open_percent = np.random.randint(10, 90)
            close_percent = 100 - open_percent       # Close percentage

            values = [open_percent, close_percent]
            colors = ['#86CD57', '#d3e7c7']  # Green for open, Light green for close

            # Create pie chart
            fig, ax = plt.subplots(figsize=(2, 2), dpi=100)
            ax.pie(values, colors=colors, startangle=90, counterclock=False, wedgeprops={'edgecolor': 'white'})

            # Convert to base64
            buffer = BytesIO()
            plt.savefig(buffer, format='png', bbox_inches='tight', transparent=True)
            buffer.seek(0)
            
            pie_charts[f'chart_{i+1}'] = base64.b64encode(buffer.getvalue()).decode()
            pie_charts_values[f'chart_{i+1}'] = {'open': open_percent, 'close': close_percent}
            buffer.close()
            plt.close(fig)  


        # Generate 3 heatmaps
        heatmaps = {}
        for _ in range(3):
            data = np.random.randint(1, 100, size=(5, 5))

            fig, ax = plt.subplots(figsize=(3, 3), dpi=100)

            cmap = plt.cm.get_cmap("RdYlGn_r")

            ax.matshow(data, cmap=cmap, alpha=0.8)

            ax.set_xticks([])
            ax.set_yticks([])

            for spine in ax.spines.values():
                spine.set_visible(False)

            buffer = BytesIO()
            plt.savefig(buffer, format='png', bbox_inches='tight', transparent=True)
            buffer.seek(0)
            heatmaps[f'chart_{_+1}'] = base64.b64encode(buffer.getvalue()).decode()

            buffer.close()
            plt.close(fig)

        context = {
            "total_risk_register": total_risk_register,
            "total_mitigated_risk": total_mitigated_risk,
            "total_unmitigated_risk": total_unmitigated_risk,
            "total_design_development_risk": total_design_development_risk,
            "total_construction_risk": total_construction_risk,
            "total_employer_change_risk": total_employer_change_risk,
            "total_employer_other_risk": total_employer_other_risk,
            "pie_charts": pie_charts,
            "pie_charts_values": pie_charts_values,
            "heatmaps": heatmaps
        }

        return self.render(context)
    
    def post(self, request, *args, **kwargs):
        if "file" not in request.FILES:
            return JsonResponse({"error": "No file uploaded"}, status=400)

        file = request.FILES["file"]

        wb = openpyxl.load_workbook(file, data_only=True)

        if "13. Risk Register" not in wb.sheetnames:
            return JsonResponse({"error": "Sheet '13. Risk Register' not found."}, status=400)

        ws = wb["13. Risk Register"]

        section_id = ""

        for row in ws.iter_rows(min_row=4, values_only=True):
            if not row[1]:  # Skip if there's no 'Ref'
                continue

            if row[1] and row[2] and all(cell is None or cell == 0 for cell in row[3:13]): # its a 'section'
                try:
                    section = models.RiskRegisterSection.objects.create(_id=str(ObjectId()), name=row[2])
                    section_id = section._id
                except Exception:
                    pass
            
            elif row[1] and all(cell is None or cell == 0 for cell in row[2:13]): # its a 'section'
                try:
                    section = models.RiskRegisterSection.objects.create(_id=str(ObjectId()), name=row[1])
                    section_id = section._id
                except Exception:
                    pass

            else:
                try:
                    models.RiskRegister.objects.create(
                        _id=str(ObjectId()),
                        ref=row[1],
                        risk_category=row[2],
                        specific_risk=row[3],
                        risk_owner=row[4],
                        likelihood=row[5] if isinstance(row[5], int) else 0,
                        impact=row[6] if isinstance(row[6], int) else 0,
                        rating=row[7] if isinstance(row[7], int) else 0,
                        mitigation=row[8],
                        mitigated_likelihood=row[9] if isinstance(row[9], int) else 0,
                        mitigated_impact=row[10] if isinstance(row[10], int) else 0,
                        mitigated_rating=row[11] if isinstance(row[11], int) else 0,
                        comments=row[12],
                        cost=row[13] if isinstance(row[13], int) else 0,
                        status=row[14],
                        section_id=section_id
                    )
                except Exception:
                    pass

        return JsonResponse({"message": "File processed successfully."}, status=201)
    

class NetCarbonDashboardView(ViewBase):
    TEMPLATE_NAME = 'dashboard/net_carbon_dashboard.html'

    def get(self, request, *args, **kwargs):

        context = {}

        return self.render(context)
    

class RetroFitDashboardView(ViewBase):
    TEMPLATE_NAME = 'dashboard/retro_fit_dashboard.html'

    def get(self, request, *args, **kwargs):

        context = {}

        return self.render(context)
    

class InformationManagementDashboardView(ViewBase):
    TEMPLATE_NAME = 'dashboard/information_management_dashboard.html'

    def get(self, request, *args, **kwargs):

        context = {}

        return self.render(context)
    

class costSummaryOperationsView(ViewBase):
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            type = data.get("type")
            section_id = data.get("section_id")
            section_name = data.get("section_name")
            row_id = data.get("row_id")
            ref = data.get("ref")
            item = data.get("item")
            contract_sum = data.get("contract_sum")
            certified_payments = data.get("certified_payments")
            accrued_payments = data.get("accrued_payments")
            total_expenditure = data.get("total_expenditure")
            variance_total = data.get("variance_total")
            variance_period = data.get("variance_period")

            if type == "section" and not section_id:
                section = models.CostSummarySection.objects.create(_id=str(ObjectId()), name=section_name)
                return JsonResponse({"message": "Cost summary saved successfully", "id": section._id}, status=201)
            
            if type == "section" and section_id:
                section = models.CostSummarySection.objects.filter(_id=section_id).update(name=section_name)
                return JsonResponse({"message": "Cost summary updated successfully", "id": section_id}, status=200)
            
            if type == "row" and not row_id and section_id:
                row = models.CostSummary.objects.create(_id=str(ObjectId()), ref=ref, item=item, contract_sum=contract_sum, certified_payments=certified_payments, accrued_payments=accrued_payments, total_expenditure=total_expenditure, variance_total=variance_total, variance_period=variance_period, section_id=section_id, created_at=timezone.now(), updated_by=request.user.get_name())
               
                cost_summary = models.CostSummary.objects.all()
                total_contract_sum = sum(Decimal(str(obj.contract_sum)) for obj in cost_summary if obj.contract_sum)
                certified_payments_sum = sum(Decimal(str(obj.certified_payments)) for obj in cost_summary if obj.certified_payments)
                anticipated_payments_sum = sum(Decimal(str(obj.accrued_payments)) for obj in cost_summary if obj.accrued_payments)
                forecast_expenditures_sum = sum(Decimal(str(obj.total_expenditure)) for obj in cost_summary if obj.total_expenditure)
                total_variance_sum = sum(Decimal(str(obj.variance_total)) for obj in cost_summary if obj.variance_total)
                total_variance_period_sum = sum(Decimal(str(obj.variance_period)) for obj in cost_summary if obj.variance_period)
                
                return JsonResponse({"message": "Cost summary row saved successfully", "id": row._id, "total_contract_sum": total_contract_sum, "certified_payments_sum": certified_payments_sum, "anticipated_payments_sum": anticipated_payments_sum, "forecast_expenditures_sum": forecast_expenditures_sum, "total_variance_sum": total_variance_sum, "total_variance_period_sum": total_variance_period_sum}, status=201)
            
            if type == "row" and row_id and section_id:
                row = models.CostSummary.objects.filter(_id=row_id).update(ref=ref, item=item, contract_sum=contract_sum, certified_payments=certified_payments, accrued_payments=accrued_payments, total_expenditure=total_expenditure, variance_total=variance_total, variance_period=variance_period, section_id=section_id, updated_at=timezone.now(), updated_by=request.user.get_name())
                
                cost_summary = models.CostSummary.objects.all()
                total_contract_sum = sum(Decimal(str(obj.contract_sum)) for obj in cost_summary if obj.contract_sum)
                certified_payments_sum = sum(Decimal(str(obj.certified_payments)) for obj in cost_summary if obj.certified_payments)
                anticipated_payments_sum = sum(Decimal(str(obj.accrued_payments)) for obj in cost_summary if obj.accrued_payments)
                forecast_expenditures_sum = sum(Decimal(str(obj.total_expenditure)) for obj in cost_summary if obj.total_expenditure)
                total_variance_sum = sum(Decimal(str(obj.variance_total)) for obj in cost_summary if obj.variance_total)
                total_variance_period_sum = sum(Decimal(str(obj.variance_period)) for obj in cost_summary if obj.variance_period)
                
                return JsonResponse({"message": "Cost summary row updated successfully", "id": row_id, "total_contract_sum": total_contract_sum, "certified_payments_sum": certified_payments_sum, "anticipated_payments_sum": anticipated_payments_sum, "forecast_expenditures_sum": forecast_expenditures_sum, "total_variance_sum": total_variance_sum, "total_variance_period_sum": total_variance_period_sum}, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    def delete(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            type = data.get("type")
            delete_row_id = request.GET.get("delete_row_id")
            delete_section_id = request.GET.get("delete_section_id")

            if type == "delete_section" and delete_section_id:
                models.CostSummarySection.objects.filter(_id=delete_section_id).delete()
                return JsonResponse({"message": "Cost summary section deleted successfully"}, status=200)

            if type == "delete_row" and delete_row_id:
                models.CostSummary.objects.filter(_id=delete_row_id).delete()

                cost_summary = models.CostSummary.objects.all()
                total_contract_sum = sum(Decimal(str(obj.contract_sum)) for obj in cost_summary if obj.contract_sum)
                certified_payments_sum = sum(Decimal(str(obj.certified_payments)) for obj in cost_summary if obj.certified_payments)
                anticipated_payments_sum = sum(Decimal(str(obj.accrued_payments)) for obj in cost_summary if obj.accrued_payments)
                forecast_expenditures_sum = sum(Decimal(str(obj.total_expenditure)) for obj in cost_summary if obj.total_expenditure)
                total_variance_sum = sum(Decimal(str(obj.variance_total)) for obj in cost_summary if obj.variance_total)
                total_variance_period_sum = sum(Decimal(str(obj.variance_period)) for obj in cost_summary if obj.variance_period)

                return JsonResponse({"message": "Cost summary row deleted successfully", "total_contract_sum": total_contract_sum, "certified_payments_sum": certified_payments_sum, "anticipated_payments_sum": anticipated_payments_sum, "forecast_expenditures_sum": forecast_expenditures_sum, "total_variance_sum": total_variance_sum, "total_variance_period_sum": total_variance_period_sum}, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
        

class contractSumOperationsView(ViewBase):
    def updateCostSummaryFields(self, section_id):
        sections = models.ContractSumSection.objects.all()
        found_index = None
        for index, section in enumerate(sections, start=1):
            if section._id == section_id:
                found_index = index
                break

        print("found_index", found_index)

        section_contract_sum = models.ContractSum.objects.filter(section_id=section_id)
        total_section_contract_sum = sum(Decimal(str(obj.contract_sum)) for obj in section_contract_sum if obj.contract_sum)
        total_section_certified_payments = sum(Decimal(str(obj.certified_payments)) for obj in section_contract_sum if obj.certified_payments)
        
        cost_summary = models.CostSummary.objects.all()[found_index - 1]
        cost_summary.contract_sum = total_section_contract_sum
        cost_summary.certified_payments = total_section_certified_payments
        cost_summary.save()

        return (total_section_contract_sum, total_section_certified_payments, cost_summary._id)

    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            type = data.get("type")
            section_id = data.get("section_id")
            section_name = data.get("section_name")
            row_id = data.get("row_id")
            ref = data.get("ref")
            item = data.get("item")
            contract_sum = data.get("contract_sum")
            certified_payments = data.get("certified_payments")

            if type == "section" and not section_id:
                section = models.ContractSumSection.objects.create(_id=str(ObjectId()), name=section_name)

                return JsonResponse({"message": "Contract sum saved successfully", "id": section._id}, status=201)
            
            if type == "section" and section_id:
                section = models.ContractSumSection.objects.filter(_id=section_id).update(name=section_name)

                return JsonResponse({"message": "Contract sum updated successfully", "id": section_id}, status=200)
            
            if type == "row" and not row_id and section_id:
                row = models.ContractSum.objects.create(_id=str(ObjectId()), ref=ref, item=item, contract_sum=contract_sum, certified_payments=certified_payments, section_id=section_id, created_at=timezone.now(), updated_by=request.user.get_name())
                total_contract_sum, total_certified_payments, summary_id = self.updateCostSummaryFields(section_id)

                cost_summary = models.CostSummary.objects.all()
                cost_total_contract_sum = sum(Decimal(str(obj.contract_sum)) for obj in cost_summary if obj.contract_sum)
                cost_total_certified_payments_sum = sum(Decimal(str(obj.certified_payments)) for obj in cost_summary if obj.certified_payments)

                return JsonResponse({"message": "Contract sum row saved successfully", "id": row._id, "total_contract_sum": total_contract_sum, "total_certified_payments": total_certified_payments, "summary_id": summary_id, "cost_total_contract_sum": cost_total_contract_sum, "cost_total_certified_payments_sum": cost_total_certified_payments_sum}, status=201)
            
            if type == "row" and row_id and section_id:
                row = models.ContractSum.objects.filter(_id=row_id).update(ref=ref, item=item, contract_sum=contract_sum, certified_payments=certified_payments, section_id=section_id, updated_at=timezone.now(), updated_by=request.user.get_name())
                total_contract_sum, total_certified_payments, summary_id = self.updateCostSummaryFields(section_id)

                cost_summary = models.CostSummary.objects.all()
                cost_total_contract_sum = sum(Decimal(str(obj.contract_sum)) for obj in cost_summary if obj.contract_sum)
                cost_total_certified_payments_sum = sum(Decimal(str(obj.certified_payments)) for obj in cost_summary if obj.certified_payments)

                return JsonResponse({"message": "Contract sum row updated successfully", "id": row_id, "total_contract_sum": total_contract_sum, "total_certified_payments": total_certified_payments, "summary_id": summary_id, "cost_total_contract_sum": cost_total_contract_sum, "cost_total_certified_payments_sum": cost_total_certified_payments_sum}, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    def delete(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            type = data.get("type")
            delete_row_id = request.GET.get("delete_row_id")
            delete_section_id = request.GET.get("delete_section_id")

            if type == "delete_section" and delete_section_id:
                models.ContractSumSection.objects.filter(_id=delete_section_id).delete()
                return JsonResponse({"message": "Cost summary section deleted successfully"}, status=200)

            if type == "delete_row" and delete_row_id:
                models.ContractSum.objects.filter(_id=delete_row_id).delete()
                return JsonResponse({"message": "Cost summary row deleted successfully"}, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
        

class changeBreakDownOperationsView(ViewBase):
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            type = data.get("type")
            section_id = data.get("section_id")
            section_name = data.get("section_name")
            row_id = data.get("row_id")
            ref = data.get("ref")
            item = data.get("item")
            certified_payments = data.get("certified_payments")
            total_expenditure = data.get("total_expenditure")
            variance_total = data.get("variance_total")
            variance_period = data.get("variance_period")

            if type == "section" and not section_id:
                section = models.ChangeBreakDownSection.objects.create(_id=str(ObjectId()), name=section_name)
                return JsonResponse({"message": "Change break down saved successfully", "id": section._id}, status=201)
            
            if type == "section" and section_id:
                section = models.ChangeBreakDownSection.objects.filter(_id=section_id).update(name=section_name)
                return JsonResponse({"message": "Change break down updated successfully", "id": section_id}, status=200)
            
            if type == "row" and not row_id and section_id:
                row = models.ChangeBreakDown.objects.create(_id=str(ObjectId()), ref=ref, item=item, certified_payments=certified_payments, total_expenditure=total_expenditure, variance_total=variance_total, variance_period=variance_period, section_id=section_id, created_at=timezone.now(), updated_by=request.user.get_name())
                return JsonResponse({"message": "Change break down row saved successfully", "id": row._id}, status=201)
            
            if type == "row" and row_id and section_id:
                row = models.ChangeBreakDown.objects.filter(_id=row_id).update(ref=ref, item=item, certified_payments=certified_payments, total_expenditure=total_expenditure, variance_total=variance_total, variance_period=variance_period, section_id=section_id, updated_at=timezone.now(), updated_by=request.user.get_name())
                return JsonResponse({"message": "Change break down row updated successfully", "id": row_id}, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

    def delete(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            type = data.get("type")
            delete_row_id = request.GET.get("delete_row_id")
            delete_section_id = request.GET.get("delete_section_id")

            if type == "delete_section" and delete_section_id:
                models.ChangeBreakDownSection.objects.filter(_id=delete_section_id).delete()
                return JsonResponse({"message": "Change break down section deleted successfully"}, status=200)

            if type == "delete_row" and delete_row_id:
                models.ChangeBreakDown.objects.filter(_id=delete_row_id).delete()
                return JsonResponse({"message": "Change break down row deleted successfully"}, status=200)

        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)
        

class getCostChartDataView(ViewBase):
    def get(self, request, *args, **kwargs):
        data = models.CostReporting.objects.values("month").distinct().order_by("month")
        months = [d["month"] for d in data if d["month"]]

        forecast_monthly_data = []
        actual_monthly_data = []
        forecast_cumulative_data = []
        actual_cumulative_data = []

        for month in months:
            forecast_monthly = models.CostReporting.objects.filter(month=month).aggregate(Sum("forecast_monthly"))["forecast_monthly__sum"] or 0
            actual_monthly = models.CostReporting.objects.filter(month=month).aggregate(Sum("actual_monthly"))["actual_monthly__sum"] or 0
            forecast_cumulative = models.CostReporting.objects.filter(month=month).aggregate(Sum("forecast_comulative"))["forecast_comulative__sum"] or 0
            actual_cumulative = models.CostReporting.objects.filter(month=month).aggregate(Sum("actual_comulative"))["actual_comulative__sum"] or 0

            forecast_monthly_data.append(forecast_monthly)
            actual_monthly_data.append(actual_monthly)
            forecast_cumulative_data.append(forecast_cumulative)
            actual_cumulative_data.append(actual_cumulative)

        months = [date.strftime("%b %y") for date in months]

        chart_data = {
            "labels": months,
            "datasets": [
                {
                    "type": "bar",
                    "label": "Forecast Monthly",
                    "data": forecast_monthly_data,
                    "backgroundColor": "rgba(133, 204, 87, 1)",
                    "borderColor": "rgba(101, 225, 193, 1)",
                    "borderWidth": 1,
                    "stack": "Stack 0"
                },
                {
                    "type": "bar",
                    "label": "Actual Monthly",
                    "data": actual_monthly_data,
                    "backgroundColor": "rgba(133, 204, 87, 1)",
                    "borderColor": "rgba(101, 225, 193, 1)",
                    "borderWidth": 1,
                    "stack": "Stack 0"
                },
                {
                    "type": "line",
                    "label": "Forecast Cumulative",
                    "data": forecast_cumulative_data,
                    "borderColor": "rgba(101, 225, 193, 1)",
                    "borderWidth": 2,
                    "fill": False
                },
                {
                    "type": "line",
                    "label": "Actual Cumulative",
                    "data": actual_cumulative_data,
                    "borderColor": "rgba(101, 225, 193, 1)",
                    "borderWidth": 2,
                    "fill": False
                }
            ]
        }
        
        return JsonResponse(chart_data)
    

class downloadCostReportView(ViewBase):

    @staticmethod
    def clear_sheet_except_header(sheet):
        max_row = sheet.max_row
        if max_row > 1:
            sheet.delete_rows(2, max_row - 1)

    def post(self, request, *args, **kwargs):
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f'cost_report_{timestamp}.xlsx'

            template_path = os.path.join(settings.BASE_DIR, 'static/cost_report_template.xlsx')
            new_file_path = os.path.join(settings.BASE_DIR, f'static/{file_name}')

            wb = openpyxl.load_workbook(template_path)

            # ==================== SHEET 1: NEW Cost Summary ====================
            cost_summary_sheet = wb["NEW Cost Summary"]
            self.clear_sheet_except_header(cost_summary_sheet)
            row = 2
            ref_counter = 1

            sections = models.CostSummarySection.objects.all()
            for section in sections:
                cost_summary_sheet.cell(row=row, column=1).value = str(ref_counter)
                cost_summary_sheet.cell(row=row, column=2).value = section.name
                cost_summary_sheet_cell = cost_summary_sheet.cell(row=row, column=2)
                cost_summary_sheet_cell.font = Font(bold=True, size=12)
                row += 1

                summaries = models.CostSummary.objects.filter(section=section)
                for summary in summaries:
                    cost_summary_sheet.cell(row=row, column=1).value = summary.ref
                    cost_summary_sheet.cell(row=row, column=2).value = summary.item
                    cost_summary_sheet.cell(row=row, column=3).value = float(summary.contract_sum or 0)
                    cost_summary_sheet.cell(row=row, column=4).value = float(summary.certified_payments or 0)
                    cost_summary_sheet.cell(row=row, column=5).value = float(summary.accrued_payments or 0)
                    cost_summary_sheet.cell(row=row, column=6).value = float(summary.total_expenditure or 0)
                    cost_summary_sheet.cell(row=row, column=7).value = float(summary.variance_total or 0)
                    cost_summary_sheet.cell(row=row, column=8).value = float(summary.variance_period or 0)
                    row += 1

                ref_counter += 1

            # ==================== SHEET 2: NEW Contract Sum ====================
            contract_sum_sheet = wb["NEW Contract Sum"]
            self.clear_sheet_except_header(contract_sum_sheet)
            row = 2
            ref_counter = 1

            contract_sections = models.ContractSumSection.objects.all()
            for section in contract_sections:
                contract_sum_sheet.cell(row=row, column=1).value = str(ref_counter)
                contract_sum_sheet.cell(row=row, column=2).value = section.name
                contract_sum_sheet_cell = contract_sum_sheet.cell(row=row, column=2)
                contract_sum_sheet_cell.font = Font(bold=True, size=12)
                row += 1

                items = models.ContractSum.objects.filter(section=section)
                for item in items:
                    contract_sum_sheet.cell(row=row, column=1).value = item.ref
                    contract_sum_sheet.cell(row=row, column=2).value = item.item
                    contract_sum_sheet.cell(row=row, column=3).value = float(item.contract_sum or 0)
                    contract_sum_sheet.cell(row=row, column=4).value = float(item.certified_payments or 0)
                    row += 1

                ref_counter += 1

            # ==================== SHEET 3: NEW EAChange ====================
            change_sheet = wb["NEW EAChange"]
            self.clear_sheet_except_header(change_sheet)
            row = 2
            ref_counter = 1

            change_sections = models.ChangeBreakDownSection.objects.all()
            for section in change_sections:
                change_sheet.cell(row=row, column=1).value = str(ref_counter)
                change_sheet.cell(row=row, column=2).value = section.name
                change_sheet_cell = change_sheet.cell(row=row, column=2)
                change_sheet_cell.font = Font(bold=True, size=12)
                row += 1

                changes = models.ChangeBreakDown.objects.filter(section=section)
                for change in changes:
                    change_sheet.cell(row=row, column=1).value = change.ref
                    change_sheet.cell(row=row, column=2).value = change.item
                    change_sheet.cell(row=row, column=3).value = float(change.certified_payments or 0)
                    change_sheet.cell(row=row, column=4).value = float(change.total_expenditure or 0)
                    change_sheet.cell(row=row, column=5).value = float(change.variance_total or 0)
                    change_sheet.cell(row=row, column=6).value = float(change.variance_period or 0)
                    row += 1

                ref_counter += 1

            wb.save(new_file_path)

            # Upload to S3
            bucket_name = 'ai-qs'
            s3_key = f"cost_reports/cost_report_{timestamp}.xlsx"

            aws_access_key_id = os.getenv("AWS_ACCESS_KEY_ID")
            aws_secret_access_key = os.getenv("AWS_SECRET_ACCESS_KEY")
            region_name = os.getenv("AWS_REGION")

            s3 = boto3.client(
                's3',
                aws_access_key_id=aws_access_key_id,
                aws_secret_access_key=aws_secret_access_key,
                region_name=region_name
            )

            s3.upload_file(
                Filename=new_file_path,
                Bucket=bucket_name,
                Key=s3_key,
                ExtraArgs={'ContentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}
            )

            s3.put_object_acl(ACL='public-read', Bucket=bucket_name, Key=s3_key)

            file_url = f"https://{bucket_name}.s3.{region_name}.amazonaws.com/{s3_key}"

            os.remove(new_file_path)

            return JsonResponse({'success': True, 'file_url': file_url, 'file_name': file_name})

        except Exception as e:
            return JsonResponse({'success': False, 'message': 'Something went wrong while exporting cost report', 'error': str(e)})